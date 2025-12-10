from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import logging

logger = logging.getLogger(__name__)
from .forms import (
    FeedbackFormForm,
    RegisterForm,
    LoginForm,
    ProfileForm,
    PasswordChangeForm,
    ApplicationForm,
    ConsultationForm,
    MeetingForm,
    ReportForm,
)
from .models import User, Application, FeedbackForm, Meeting, Report, Notification


def get_breadcrumbs(items):
    """Вспомогательная функция для создания breadcrumbs"""
    breadcrumbs = [{"title": "Главная", "url": reverse("index")}]
    breadcrumbs.extend(items)
    return breadcrumbs


def create_notification(user, notification_type, title, message, application=None, meeting=None):
    """Создание уведомления для пользователя"""
    Notification.objects.create(
        user=user,
        notification_type=notification_type,
        title=title,
        message=message,
        related_application=application,
        related_meeting=meeting,
    )


def index(request):
    return render(request, "index.html", {"breadcrumbs": []})


def about(request):
    breadcrumbs = get_breadcrumbs([{"title": "О нас", "url": ""}])
    return render(request, "about.html", {"breadcrumbs": breadcrumbs})


def faq(request):
    breadcrumbs = get_breadcrumbs([{"title": "FAQ", "url": ""}])
    return render(request, "faq.html", {"breadcrumbs": breadcrumbs})


def contacts(request):
    breadcrumbs = get_breadcrumbs([{"title": "Контакты", "url": ""}])
    return render(request, "contacts.html", {"breadcrumbs": breadcrumbs})


def privacy(request):
    breadcrumbs = get_breadcrumbs([{"title": "Конфиденциальность", "url": ""}])
    return render(request, "privacy.html", {"breadcrumbs": breadcrumbs})


def services(request):
    breadcrumbs = get_breadcrumbs([{"title": "Услуги", "url": ""}])
    return render(request, "services.html", {"breadcrumbs": breadcrumbs})


def support(request):
    breadcrumbs = get_breadcrumbs([{"title": "Мотивация", "url": ""}])
    return render(request, "support.html", {"breadcrumbs": breadcrumbs})


def feedback(request):
    if request.method == "POST":
        form = FeedbackFormForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                "Спасибо за ваше сообщение! Мы свяжемся с вами в ближайшее время.",
            )
            return redirect("feedback")
    else:
        form = FeedbackFormForm()
    breadcrumbs = get_breadcrumbs([{"title": "Обратная связь", "url": ""}])
    return render(request, "feedback.html", {"form": form, "breadcrumbs": breadcrumbs})


def register_view(request):
    if request.user.is_authenticated:
        return redirect("index")

    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect("index")
    else:
        form = RegisterForm()
    breadcrumbs = get_breadcrumbs([{"title": "Регистрация", "url": ""}])
    return render(request, "register.html", {"form": form, "breadcrumbs": breadcrumbs})


def login_view(request):
    if request.user.is_authenticated:
        return redirect("index")

    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get("username")
            password = form.cleaned_data.get("password")
            from django.contrib.auth import authenticate

            user = authenticate(username=username, password=password)
            if user:
                login(request, user)
                next_url = request.GET.get("next", "index")
                return redirect(next_url)
    else:
        form = LoginForm()
    breadcrumbs = get_breadcrumbs([{"title": "Вход", "url": ""}])
    return render(request, "login.html", {"form": form, "breadcrumbs": breadcrumbs})


@login_required
def logout_view(request):
    logout(request)
    return redirect("index")


@login_required
def profile_view(request):
    if request.method == "POST":
        if "update_profile" in request.POST:
            profile_form = ProfileForm(request.POST, instance=request.user)
            password_form = PasswordChangeForm(user=request.user)
            if profile_form.is_valid():
                try:
                    profile_form.save()
                    messages.success(request, "Профиль успешно обновлен.")
                    return redirect("profile")
                except ValidationError as e:
                    logger.error(f"Validation error in profile update: {e}")
                    messages.error(
                        request, f"Ошибка валидации: {', '.join(e.messages)}"
                    )
                except Exception as e:
                    logger.error(f"Error updating profile: {str(e)}", exc_info=True)
                    messages.error(
                        request,
                        "Произошла ошибка при обновлении профиля. Пожалуйста, попробуйте еще раз.",
                    )
            else:
                messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
        elif "change_password" in request.POST:
            profile_form = ProfileForm(instance=request.user)
            password_form = PasswordChangeForm(user=request.user, data=request.POST)
            if password_form.is_valid():
                try:
                    user = password_form.save()
                    update_session_auth_hash(request, user)
                    messages.success(request, "Пароль успешно изменен.")
                    return redirect("profile")
                except ValidationError as e:
                    logger.error(f"Validation error in password change: {e}")
                    messages.error(
                        request, f"Ошибка валидации: {', '.join(e.messages)}"
                    )
                except Exception as e:
                    logger.error(f"Error changing password: {str(e)}", exc_info=True)
                    messages.error(
                        request,
                        "Произошла ошибка при смене пароля. Пожалуйста, попробуйте еще раз.",
                    )
            else:
                messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
        else:
            profile_form = ProfileForm(instance=request.user)
            password_form = PasswordChangeForm(user=request.user)
    else:
        profile_form = ProfileForm(instance=request.user)
        password_form = PasswordChangeForm(user=request.user)

    breadcrumbs = get_breadcrumbs([{"title": "Профиль", "url": ""}])
    return render(
        request,
        "profile.html",
        {
            "profile_form": profile_form,
            "password_form": password_form,
            "breadcrumbs": breadcrumbs,
        },
    )


@login_required
def create_application(request, psychologist_id=None):
    """Создание заявки психологу"""
    if request.user.is_psychologist():
        messages.error(
            request,
            "Психологи не могут подавать заявки. Используйте панель управления для работы с заявками.",
        )
        return redirect("index")

    if request.method == "POST":
        form = ApplicationForm(request.POST)
        if form.is_valid():
            application = form.save(commit=False)
            application.user = request.user
            application.status = "pending"
            application.save()
            if application.psychologist:
                create_notification(
                    user=application.psychologist,
                    notification_type="application_assigned",
                    title="Новая заявка",
                    message=f"Студент {request.user.get_display_name()} подал(а) вам новую заявку: '{application.title}'",
                    application=application,
                )
            messages.success(
                request,
                "Заявка успешно отправлена! Психолог рассмотрит её в ближайшее время.",
            )
            return redirect("my_applications")
    else:
        form = ApplicationForm()
        if psychologist_id:
            psychologist = get_object_or_404(
                User, id=psychologist_id, role="psychologist"
            )
            form.fields["psychologist"].initial = psychologist

    breadcrumbs = get_breadcrumbs([{"title": "Подать заявку", "url": ""}])
    return render(
        request, "application_create.html", {"form": form, "breadcrumbs": breadcrumbs}
    )


@login_required
def my_applications(request):
    """Список заявок текущего пользователя"""
    applications = Application.objects.filter(user=request.user).order_by("-created_at")
    breadcrumbs = get_breadcrumbs([{"title": "Мои заявки", "url": ""}])
    return render(
        request,
        "my_applications.html",
        {"applications": applications, "breadcrumbs": breadcrumbs},
    )


def psychologists_list(request):
    """Список всех психологов"""
    psychologists = User.objects.filter(role="psychologist").order_by("username")
    breadcrumbs = get_breadcrumbs([{"title": "Психологи", "url": ""}])
    return render(
        request,
        "psychologists_list.html",
        {"psychologists": psychologists, "breadcrumbs": breadcrumbs},
    )


@login_required
def dashboard_student(request):
    """Панель управления для студентов"""
    if request.user.is_admin_user():
        return redirect("dashboard_admin")
    if request.user.is_psychologist():
        return redirect("dashboard_psychologist")

    applications = Application.objects.filter(user=request.user)

    status_filter = request.GET.get("status")
    if status_filter:
        applications = applications.filter(status=status_filter)

    applications = applications.order_by("-created_at")
    meetings = Meeting.objects.filter(student=request.user).order_by("date", "time")

    breadcrumbs = get_breadcrumbs([{"title": "Панель управления", "url": ""}])
    return render(
        request,
        "dashboard_student.html",
        {
            "applications": applications,
            "meetings": meetings,
            "current_status": status_filter,
            "breadcrumbs": breadcrumbs,
        },
    )


@login_required
def dashboard_admin(request):
    """Панель управления для администраторов"""
    if not request.user.is_admin_user():
        return redirect("index")

    if request.method == "POST" and "change_role" in request.POST:
        user_id = request.POST.get("user_id")
        new_role = request.POST.get("role")
        if user_id and new_role:
            user = get_object_or_404(User, id=user_id)
            user.role = new_role
            user.save()
            messages.success(
                request,
                f"Роль пользователя {user.get_display_name()} изменена на {user.get_role_display()}",
            )
            return redirect("dashboard_admin")

    users = User.objects.all()
    user_role_filter = request.GET.get("user_role")
    if user_role_filter:
        users = users.filter(role=user_role_filter)
    users = users.order_by("-created_at")

    feedback_forms = FeedbackForm.objects.all()
    feedback_status_filter = request.GET.get("feedback_status")
    if feedback_status_filter:
        feedback_forms = feedback_forms.filter(status=feedback_status_filter)
    feedback_forms = feedback_forms.order_by("-created_at")

    meetings = Meeting.objects.all()
    meeting_date_filter = request.GET.get("meeting_date")
    if meeting_date_filter:
        meetings = meetings.filter(date=meeting_date_filter)
    meetings = meetings.order_by("date", "time")

    reports = Report.objects.all().order_by("-created_at")

    breadcrumbs = get_breadcrumbs(
        [{"title": "Панель управления администратора", "url": ""}]
    )
    return render(
        request,
        "dashboard_admin.html",
        {
            "users": users,
            "feedback_forms": feedback_forms,
            "meetings": meetings,
            "reports": reports,
            "current_user_role": user_role_filter,
            "current_feedback_status": feedback_status_filter,
            "current_meeting_date": meeting_date_filter,
            "breadcrumbs": breadcrumbs,
        },
    )


@login_required
def dashboard_psychologist(request):
    """Панель управления для психологов"""
    if not request.user.is_psychologist():
        return redirect("index")

    applications = Application.objects.filter(psychologist=request.user)

    status_filter = request.GET.get("status")
    if status_filter:
        applications = applications.filter(status=status_filter)

    applications = applications.order_by("-created_at")

    general_applications = Application.objects.filter(
        psychologist__isnull=True, status="pending"
    ).order_by("-created_at")

    meetings = Meeting.objects.filter(psychologist=request.user).order_by(
        "date", "time"
    )

    reports = Report.objects.filter(psychologist=request.user).order_by("-created_at")

    breadcrumbs = get_breadcrumbs([{"title": "Панель управления психолога", "url": ""}])
    return render(
        request,
        "dashboard_psychologist.html",
        {
            "applications": applications,
            "general_applications": general_applications,
            "meetings": meetings,
            "reports": reports,
            "current_status": status_filter,
            "breadcrumbs": breadcrumbs,
        },
    )


@login_required
def application_detail_psychologist(request, application_id):
    """Детальный просмотр заявки для психолога и администратора"""
    if not request.user.is_psychologist() and not request.user.is_admin_user():
        return redirect("index")

    application = get_object_or_404(Application, id=application_id)

    # Администраторы могут видеть все заявки, психологи - только свои
    if not request.user.is_admin_user():
        if application.psychologist and application.psychologist != request.user:
            messages.error(request, "У вас нет доступа к этой заявке.")
            return redirect("dashboard_psychologist")

    if request.method == "POST":
        if "take_application" in request.POST:
            if not application.psychologist:
                application.psychologist = request.user
                application.status = "in_progress"
                application.save()
                create_notification(
                    user=request.user,
                    notification_type="application_assigned",
                    title="Заявка назначена",
                    message=f"Вам назначена заявка #{application.id} '{application.title}' от студента {application.user.get_display_name()}",
                    application=application,
                )
                messages.success(request, "Заявка успешно взята в работу.")
                return redirect(
                    "application_detail_psychologist", application_id=application.id
                )
        elif "add_consultation" in request.POST:
            if application.status == "completed":
                messages.error(
                    request, "Нельзя отправлять ответы для завершенной заявки."
                )
                return redirect(
                    "application_detail_psychologist", application_id=application.id
                )
            consultation_form = ConsultationForm(request.POST)
            if consultation_form.is_valid():
                consultation = consultation_form.save(commit=False)
                consultation.application = application
                consultation.psychologist = request.user
                consultation.save()
                if application.status == "pending":
                    application.status = "in_progress"
                    application.psychologist = request.user
                    application.save()
                create_notification(
                    user=application.user,
                    notification_type="consultation_message",
                    title=f"Новое сообщение от психолога",
                    message=f"Психолог {request.user.get_display_name()} отправил(а) вам сообщение по заявке #{application.id}",
                    application=application,
                )
                messages.success(request, "Ответ успешно отправлен студенту.")
                return redirect(
                    "application_detail_psychologist", application_id=application.id
                )
            else:
                messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
        elif "add_meeting" in request.POST:
            if application.status == "completed":
                messages.error(
                    request, "Нельзя назначать встречи для завершенной заявки."
                )
                return redirect(
                    "application_detail_psychologist", application_id=application.id
                )
            meeting_form = MeetingForm(request.POST, psychologist=request.user)
            consultation_form = ConsultationForm()
            report_form = ReportForm()
            if meeting_form.is_valid():
                meeting = meeting_form.save(commit=False)
                meeting.student = application.user
                meeting.psychologist = request.user
                meeting.application = application
                meeting.save()
                create_notification(
                    user=application.user,
                    notification_type="meeting_scheduled",
                    title="Встреча назначена",
                    message=f"Психолог {request.user.get_display_name()} назначил(а) вам встречу на {meeting.date.strftime('%d.%m.%Y')} в {meeting.time.strftime('%H:%M')}",
                    application=application,
                    meeting=meeting,
                )
                messages.success(request, "Встреча успешно назначена.")
                return redirect(
                    "application_detail_psychologist", application_id=application.id
                )
            else:
                errors_attr = getattr(meeting_form, "non_field_errors", None)
                if errors_attr is not None:
                    if callable(errors_attr):
                        errors = errors_attr()
                    else:
                        errors = errors_attr

                    if errors:
                        for error in errors:
                            messages.error(request, str(error))
                    else:
                        messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
                else:
                    messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
                consultation_form = ConsultationForm()
                report_form = ReportForm()
        elif "complete_application" in request.POST:
            if application.psychologist == request.user:
                report_form = ReportForm(request.POST)
                consultation_form = ConsultationForm()
                meeting_form = MeetingForm(psychologist=request.user)
                if report_form.is_valid():
                    report = report_form.save(commit=False)
                    report.application = application
                    report.psychologist = request.user
                    report.student = application.user
                    report.save()
                    application.status = "completed"
                    application.save()
                    create_notification(
                        user=application.user,
                        notification_type="application_completed",
                        title="Заявка завершена",
                        message=f"Ваша заявка #{application.id} '{application.title}' была завершена психологом {request.user.get_display_name()}",
                        application=application,
                    )
                    messages.success(request, "Заявка успешно завершена и отчет добавлен.")
                    return redirect(
                        "application_detail_psychologist", application_id=application.id
                    )
                else:
                    messages.error(request, "Пожалуйста, заполните отчет корректно.")
                    consultation_form = ConsultationForm()
                    meeting_form = MeetingForm(psychologist=request.user)
                    # report_form уже определен выше, не нужно переопределять
            else:
                consultation_form = ConsultationForm()
                meeting_form = MeetingForm(psychologist=request.user)
                report_form = ReportForm()
        else:
            consultation_form = ConsultationForm()
            meeting_form = MeetingForm(psychologist=request.user)
            report_form = ReportForm()
    else:
        consultation_form = ConsultationForm()
        meeting_form = MeetingForm(psychologist=request.user)
        report_form = ReportForm()

    consultations = application.consultations.all().order_by("-created_at")

    # Определяем правильный URL для breadcrumbs в зависимости от роли
    if request.user.is_admin_user():
        dashboard_url = reverse("dashboard_admin")
        dashboard_title = "Панель управления администратора"
    else:
        dashboard_url = reverse("dashboard_psychologist")
        dashboard_title = "Панель управления"
    
    breadcrumbs = get_breadcrumbs(
        [
            {"title": dashboard_title, "url": dashboard_url},
            {"title": f"Заявка #{application.id}", "url": ""},
        ]
    )
    # Проверяем, есть ли уже отчет для этой заявки
    existing_report = Report.objects.filter(application=application).first()
    
    return render(
        request,
        "application_detail_psychologist.html",
        {
            "application": application,
            "consultations": consultations,
            "consultation_form": consultation_form,
            "meeting_form": meeting_form,
            "report_form": report_form,
            "existing_report": existing_report,
            "breadcrumbs": breadcrumbs,
        },
    )


@login_required
def application_detail_student(request, application_id):
    """Детальный просмотр заявки для студента"""
    if request.user.is_psychologist() or request.user.is_admin_user():
        return redirect("index")

    application = get_object_or_404(Application, id=application_id)

    if application.user != request.user:
        messages.error(request, "У вас нет доступа к этой заявке.")
        return redirect("dashboard")

    if request.method == "POST" and "add_message" in request.POST:
        if application.status == "completed":
            messages.error(
                request, "Нельзя отправлять сообщения для завершенной заявки."
            )
            return redirect(
                "application_detail_student", application_id=application.id
            )
        consultation_form = ConsultationForm(request.POST)
        if consultation_form.is_valid():
            consultation = consultation_form.save(commit=False)
            consultation.application = application
            consultation.student = request.user
            consultation.save()
            if application.psychologist:
                create_notification(
                    user=application.psychologist,
                    notification_type="consultation_message",
                    title=f"Новое сообщение от студента",
                    message=f"Студент {request.user.get_display_name()} отправил(а) вам сообщение по заявке #{application.id}",
                    application=application,
                )
            messages.success(request, "Сообщение успешно отправлено.")
            return redirect(
                "application_detail_student", application_id=application.id
            )
        else:
            messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
    else:
        consultation_form = ConsultationForm()

    consultations = application.consultations.all().order_by("-created_at")

    breadcrumbs = get_breadcrumbs(
        [
            {"title": "Панель управления", "url": reverse("dashboard")},
            {"title": f"Заявка #{application.id}", "url": ""},
        ]
    )
    return render(
        request,
        "application_detail_student.html",
        {
            "application": application,
            "consultations": consultations,
            "consultation_form": consultation_form,
            "breadcrumbs": breadcrumbs,
        },
    )


@login_required
def feedback_detail_admin(request, feedback_id):
    """Детальный просмотр формы обратной связи для администратора"""
    if not request.user.is_admin_user():
        return redirect("index")

    feedback = get_object_or_404(FeedbackForm, id=feedback_id)

    if request.method == "POST" and "change_status" in request.POST:
        new_status = request.POST.get("status")
        if new_status:
            feedback.status = new_status
            feedback.save()
            messages.success(request, "Статус формы обратной связи обновлен.")
            return redirect("feedback_detail_admin", feedback_id=feedback.id)

    breadcrumbs = get_breadcrumbs(
        [
            {"title": "Панель управления", "url": reverse("dashboard_admin")},
            {"title": f"Форма обратной связи #{feedback.id}", "url": ""},
        ]
    )
    return render(
        request,
        "feedback_detail_admin.html",
        {
            "feedback": feedback,
            "breadcrumbs": breadcrumbs,
        },
    )


@login_required
def delete_user(request, user_id):
    """Удаление пользователя администратором"""
    if not request.user.is_admin_user():
        messages.error(request, "У вас нет прав для выполнения этого действия.")
        return redirect("index")

    user_to_delete = get_object_or_404(User, id=user_id)

    if user_to_delete.id == request.user.id:
        messages.error(request, "Вы не можете удалить свой собственный аккаунт.")
        return redirect("dashboard_admin")

    admin_count = User.objects.filter(role="admin").count()
    if user_to_delete.role == "admin" and admin_count <= 1:
        messages.error(request, "Нельзя удалить последнего администратора.")
        return redirect("dashboard_admin")

    if request.method == "POST":
        try:
            user_display_name = user_to_delete.get_display_name()
            user_to_delete.delete()
            messages.success(
                request, f"Пользователь {user_display_name} успешно удален."
            )
            return redirect("dashboard_admin")
        except Exception as e:
            logger.error(f"Error deleting user: {str(e)}", exc_info=True)
            messages.error(request, "Произошла ошибка при удалении пользователя.")
            return redirect("dashboard_admin")

    breadcrumbs = get_breadcrumbs(
        [
            {"title": "Панель управления", "url": reverse("dashboard_admin")},
            {"title": "Удаление пользователя", "url": ""},
        ]
    )
    return render(
        request,
        "delete_user_confirm.html",
        {
            "user_to_delete": user_to_delete,
            "breadcrumbs": breadcrumbs,
        },
    )


@login_required
def export_meetings_excel(request):
    """Экспорт встреч в Excel файл для администраторов"""
    if not request.user.is_admin_user():
        return redirect("index")

    meetings = Meeting.objects.all().select_related(
        "student", "psychologist", "application"
    )

    meeting_date_filter = request.GET.get("meeting_date")
    if meeting_date_filter:
        meetings = meetings.filter(date=meeting_date_filter)

    meetings = meetings.order_by("date", "time")

    wb = Workbook()
    ws = wb.active
    ws.title = "Встречи"

    header_fill = PatternFill(
        start_color="667eea", end_color="667eea", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)

    headers = [
        "ID",
        "Студент",
        "Email студента",
        "Психолог",
        "Email психолога",
        "Дата встречи",
        "Время встречи",
        "ID заявки",
        "Дата создания",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num, meeting in enumerate(meetings, 2):
        psychologist_name = meeting.psychologist.get_display_name()
        student_name = meeting.student.get_display_name()

        ws.cell(row=row_num, column=1, value=meeting.id)
        ws.cell(row=row_num, column=2, value=student_name)
        ws.cell(row=row_num, column=3, value=meeting.student.email)
        ws.cell(row=row_num, column=4, value=psychologist_name)
        ws.cell(row=row_num, column=5, value=meeting.psychologist.email)
        ws.cell(row=row_num, column=6, value=meeting.date.strftime("%d.%m.%Y"))
        ws.cell(row=row_num, column=7, value=meeting.time.strftime("%H:%M"))
        ws.cell(
            row=row_num,
            column=8,
            value=meeting.application.id if meeting.application else "",
        )
        ws.cell(
            row=row_num, column=9, value=meeting.created_at.strftime("%d.%m.%Y %H:%M")
        )

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"meetings_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
def edit_meeting(request, meeting_id):
    """Редактирование встречи психологом"""
    if not request.user.is_psychologist():
        messages.error(request, "У вас нет прав для выполнения этого действия.")
        return redirect("index")

    meeting = get_object_or_404(Meeting, id=meeting_id)

    if meeting.psychologist != request.user:
        messages.error(request, "У вас нет доступа к этой встрече.")
        return redirect("dashboard_psychologist")

    if request.method == "POST":
        meeting_form = MeetingForm(
            request.POST, psychologist=request.user, instance=meeting
        )
        if meeting_form.is_valid():
            meeting_form.save()
            meeting.refresh_from_db()
            create_notification(
                user=meeting.student,
                notification_type="meeting_updated",
                title="Встреча изменена",
                message=f"Психолог {request.user.get_display_name()} изменил(а) время встречи на {meeting.date.strftime('%d.%m.%Y')} в {meeting.time.strftime('%H:%M')}",
                meeting=meeting,
            )
            messages.success(request, "Встреча успешно обновлена.")
            return redirect("dashboard_psychologist")
        else:
            errors_attr = getattr(meeting_form, "non_field_errors", None)
            if errors_attr is not None:
                if callable(errors_attr):
                    errors = errors_attr()
                else:
                    errors = errors_attr

                if errors:
                    for error in errors:
                        messages.error(request, str(error))
                else:
                    messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
            else:
                messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
    else:
        meeting_form = MeetingForm(psychologist=request.user, instance=meeting)

    breadcrumbs = get_breadcrumbs(
        [
            {"title": "Панель управления", "url": reverse("dashboard_psychologist")},
            {"title": "Редактирование встречи", "url": ""},
        ]
    )
    return render(
        request,
        "edit_meeting.html",
        {
            "meeting": meeting,
            "meeting_form": meeting_form,
            "breadcrumbs": breadcrumbs,
        },
    )


@login_required
def delete_meeting(request, meeting_id):
    """Удаление встречи психологом"""
    if not request.user.is_psychologist():
        messages.error(request, "У вас нет прав для выполнения этого действия.")
        return redirect("index")

    meeting = get_object_or_404(Meeting, id=meeting_id)

    if meeting.psychologist != request.user:
        messages.error(request, "У вас нет доступа к этой встрече.")
        return redirect("dashboard_psychologist")

    if request.method == "POST":
        try:
            student_name = meeting.student.get_display_name()
            student = meeting.student
            meeting_date = meeting.date.strftime("%d.%m.%Y")
            meeting_time = meeting.time.strftime("%H:%M")
            meeting.delete()
            create_notification(
                user=student,
                notification_type="meeting_cancelled",
                title="Встреча отменена",
                message=f"Встреча на {meeting_date} в {meeting_time} была отменена",
            )
            messages.success(
                request,
                f"Встреча со студентом {student_name} на {meeting_date} в {meeting_time} успешно удалена.",
            )
            return redirect("dashboard_psychologist")
        except Exception as e:
            logger.error(f"Error deleting meeting: {str(e)}", exc_info=True)
            messages.error(request, "Произошла ошибка при удалении встречи.")
            return redirect("dashboard_psychologist")

    breadcrumbs = get_breadcrumbs(
        [
            {"title": "Панель управления", "url": reverse("dashboard_psychologist")},
            {"title": "Удаление встречи", "url": ""},
        ]
    )
    return render(
        request,
        "delete_meeting_confirm.html",
        {
            "meeting": meeting,
            "breadcrumbs": breadcrumbs,
        },
    )


@login_required
def notifications_list(request):
    """Получение списка уведомлений для текущего пользователя"""
    notifications = Notification.objects.filter(user=request.user).order_by("-created_at")[:10]
    unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
    
    from django.http import JsonResponse
    return JsonResponse({
        "notifications": [
            {
                "id": n.id,
                "title": n.title,
                "message": n.message,
                "type": n.notification_type,
                "is_read": n.is_read,
                "created_at": n.created_at.strftime("%d.%m.%Y %H:%M"),
                "application_id": n.related_application.id if n.related_application else None,
            }
            for n in notifications
        ],
        "unread_count": unread_count,
    })


@login_required
def mark_notification_read(request, notification_id):
    """Отметить уведомление как прочитанное"""
    if request.method != "POST":
        from django.http import JsonResponse
        return JsonResponse({"error": "Method not allowed"}, status=405)
    
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.is_read = True
    notification.save()
    
    from django.http import JsonResponse
    return JsonResponse({"success": True})


@login_required
def mark_all_notifications_read(request):
    """Отметить все уведомления как прочитанные"""
    if request.method != "POST":
        from django.http import JsonResponse
        return JsonResponse({"error": "Method not allowed"}, status=405)
    
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    
    from django.http import JsonResponse
    return JsonResponse({"success": True})
